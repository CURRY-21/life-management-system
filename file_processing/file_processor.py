import os
import tempfile
from typing import Dict, Any
import subprocess

class FileProcessor:
    """文件处理器类，支持多种文件格式的解析和内容提取"""
    
    def __init__(self):
        """初始化文件处理器，尝试导入所需库"""
        self.import_libraries()
    
    def import_libraries(self):
        """动态导入所需库，避免安装问题"""
        self.libraries = {
            'docx': None,
            'openpyxl': None,
            'pandas': None,
            'PIL': None,
            'speech_recognition': None,
            'PyPDF2': None,
            'cv2': None,
            'moviepy': None
        }
        
        # 尝试导入各个库
        try:
            import docx
            self.libraries['docx'] = docx
        except ImportError:
            pass
        
        try:
            import openpyxl
            self.libraries['openpyxl'] = openpyxl
        except ImportError:
            pass
        
        try:
            import pandas as pd
            self.libraries['pandas'] = pd
        except ImportError:
            pass
        
        try:
            from PIL import Image
            self.libraries['PIL'] = Image
        except ImportError:
            pass
        
        try:
            import speech_recognition as sr
            self.libraries['speech_recognition'] = sr
        except ImportError:
            pass
        
        try:
            import PyPDF2
            self.libraries['PyPDF2'] = PyPDF2
        except ImportError:
            pass
        
        try:
            import cv2
            self.libraries['cv2'] = cv2
        except ImportError:
            pass
        
        try:
            import moviepy.editor as mp
            self.libraries['moviepy'] = mp
        except ImportError:
            pass
    
    def process_file(self, file_path: str) -> str:
        """根据文件扩展名选择合适的处理方法"""
        _, ext = os.path.splitext(file_path.lower())
        
        if ext in ['.txt', '.md', '.csv']:
            return self.process_text_file(file_path)
        elif ext in ['.docx', '.wps']:
            return self.process_word_file(file_path)
        elif ext in ['.xlsx', '.xls', '.et']:
            return self.process_excel_file(file_path)
        elif ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
            return self.process_image_file(file_path)
        elif ext in ['.mp3', '.wav', '.ogg', '.flac', '.m4a']:
            return self.process_audio_file(file_path)
        elif ext in ['.mp4', '.avi', '.mov', '.wmv', '.flv']:
            return self.process_video_file(file_path)
        elif ext in ['.pdf']:
            return self.process_pdf_file(file_path)
        elif ext in ['.dps']:
            return self.process_wps_presentation(file_path)
        else:
            return f"不支持的文件格式: {ext}"
    
    def process_text_file(self, file_path: str) -> str:
        """处理文本文件"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()
        except Exception as e:
            return f"处理文本文件时出错: {str(e)}"
    
    def process_word_file(self, file_path: str) -> str:
        """处理Word文件，增强对WPS文件的支持"""
        if not self.libraries['docx']:
            return "未安装python-docx库，无法处理Word文件"
        
        try:
            # 尝试使用python-docx库打开文件
            doc = self.libraries['docx'].Document(file_path)
            full_text = []
            for para in doc.paragraphs:
                full_text.append(para.text)
            return '\n'.join(full_text)
        except Exception as e:
            # 处理WPS文件兼容性问题
            if 'PackageNotFoundError' in str(e) or 'docx.opc.exceptions.PackageNotFoundError' in str(e):
                # 尝试使用python-docx的替代方法处理WPS文件
                try:
                    from zipfile import ZipFile
                    import xml.etree.ElementTree as ET
                    
                    # 直接解析DOCX文件结构（WPS生成的DOCX有时结构略有不同）
                    full_text = []
                    with ZipFile(file_path, 'r') as zip_ref:
                        # 读取文档内容XML
                        if 'word/document.xml' in zip_ref.namelist():
                            with zip_ref.open('word/document.xml') as xml_file:
                                tree = ET.parse(xml_file)
                                root = tree.getroot()
                                
                                # 定义命名空间
                                ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
                                
                                # 查找所有段落
                                paragraphs = root.findall('.//w:p', ns)
                                for para in paragraphs:
                                    # 查找段落中的所有文本
                                    texts = para.findall('.//w:t', ns)
                                    para_text = ''.join([t.text if t.text else '' for t in texts])
                                    if para_text.strip():
                                        full_text.append(para_text)
                    
                    if full_text:
                        return '\n'.join(full_text)
                    else:
                        return f"成功解析WPS DOCX文件，但未提取到文本内容。文件：{file_path}"
                except Exception as inner_e:
                    return f"处理WPS DOCX文件时出错：{str(inner_e)}。您可以尝试将文件另存为标准DOCX格式后重新上传。"
            else:
                return f"处理Word文件时出错：{str(e)}"
    
    def process_excel_file(self, file_path: str) -> str:
        """处理Excel文件，增强对WPS文件的支持"""
        try:
            if self.libraries['pandas']:
                # 使用pandas处理，支持多种格式，包括WPS生成的Excel文件
                try:
                    # 尝试使用默认引擎
                    df = self.libraries['pandas'].read_excel(file_path)
                    return df.to_string()
                except Exception as pandas_e:
                    # 尝试使用不同的引擎处理WPS文件
                    try:
                        df = self.libraries['pandas'].read_excel(file_path, engine='openpyxl')
                        return df.to_string()
                    except Exception as engine_e:
                        # 如果pandas失败，尝试使用openpyxl直接处理
                        if self.libraries['openpyxl']:
                            return self._process_excel_with_openpyxl(file_path)
                        else:
                            return f"使用pandas处理Excel文件时出错：{str(pandas_e)}。未安装openpyxl作为备选方案。"
            elif self.libraries['openpyxl']:
                # 直接使用openpyxl处理
                return self._process_excel_with_openpyxl(file_path)
            else:
                return "未安装pandas或openpyxl库，无法处理Excel文件"
        except Exception as e:
            return f"处理Excel文件时出错：{str(e)}。如果是WPS生成的文件，尝试将其另存为标准Excel格式后重新上传。"
    
    def _process_excel_with_openpyxl(self, file_path: str) -> str:
        """使用openpyxl处理Excel文件的辅助方法"""
        try:
            # 使用openpyxl处理，添加WPS兼容性选项
            wb = self.libraries['openpyxl'].load_workbook(
                file_path, 
                read_only=True, 
                data_only=True,  # 读取单元格的计算结果而非公式
                keep_links=False  # 忽略外部链接，增强WPS兼容性
            )
            full_text = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                full_text.append(f"Sheet: {sheet}")
                for row in ws.iter_rows(values_only=True):
                    row_text = '\t'.join([str(cell) if cell is not None else '' for cell in row])
                    full_text.append(row_text)
            return '\n'.join(full_text)
        except Exception as e:
            return f"使用openpyxl处理Excel文件时出错：{str(e)}"
    
    def process_image_file(self, file_path: str) -> str:
        """处理图片文件，使用OCR提取文字"""
        if not self.libraries['PIL']:
            return "未安装PIL库，无法处理图片文件"
        
        try:
            img = self.libraries['PIL'].open(file_path)
            # 尝试使用OCR提取文字
            try:
                import pytesseract
                text = pytesseract.image_to_string(img, lang='chi_sim+eng')
                return text
            except ImportError:
                return f"图片文件：{os.path.basename(file_path)}。未安装pytesseract库，无法提取文字内容。"
            except Exception as ocr_e:
                return f"图片文件：{os.path.basename(file_path)}。OCR提取文字失败：{str(ocr_e)}"
        except Exception as e:
            return f"处理图片文件时出错: {str(e)}"
    
    def process_audio_file(self, file_path: str) -> str:
        """处理音频文件，使用语音识别提取文字"""
        if not self.libraries['speech_recognition']:
            return "未安装speech_recognition库，无法处理音频文件"
        
        try:
            r = self.libraries['speech_recognition'].Recognizer()
            with self.libraries['speech_recognition'].AudioFile(file_path) as source:
                audio_data = r.record(source)
                # 尝试使用Google语音识别
                text = r.recognize_google(audio_data, language='zh-CN')
                return text
        except Exception as e:
            return f"处理音频文件时出错: {str(e)}"
    
    def process_video_file(self, file_path: str) -> str:
        """处理视频文件，提取关键帧和音频"""
        try:
            video_info = []
            video_info.append(f"视频文件：{os.path.basename(file_path)}")
            video_info.append(f"文件大小：{os.path.getsize(file_path) / (1024 * 1024):.2f} MB")
            
            # 尝试使用OpenCV提取视频信息
            if self.libraries['cv2']:
                try:
                    cap = self.libraries['cv2'].VideoCapture(file_path)
                    if cap.isOpened():
                        fps = cap.get(self.libraries['cv2'].CAP_PROP_FPS)
                        frame_count = int(cap.get(self.libraries['cv2'].CAP_PROP_FRAME_COUNT))
                        duration = frame_count / fps if fps > 0 else 0
                        width = int(cap.get(self.libraries['cv2'].CAP_PROP_FRAME_WIDTH))
                        height = int(cap.get(self.libraries['cv2'].CAP_PROP_FRAME_HEIGHT))
                        
                        video_info.append(f"分辨率：{width}x{height}")
                        video_info.append(f"帧率：{fps:.2f}")
                        video_info.append(f"总帧数：{frame_count}")
                        video_info.append(f"时长：{duration:.2f} 秒")
                        
                        cap.release()
                except Exception as cv2_e:
                    video_info.append(f"提取视频信息失败：{str(cv2_e)}")
            
            # 尝试使用moviepy提取音频
            if self.libraries['moviepy']:
                try:
                    video = self.libraries['moviepy'].VideoFileClip(file_path)
                    audio = video.audio
                    if audio:
                        video_info.append("视频包含音频轨道")
                        audio_duration = audio.duration
                        video_info.append(f"音频时长：{audio_duration:.2f} 秒")
                except Exception as moviepy_e:
                    video_info.append(f"提取音频信息失败：{str(moviepy_e)}")
            
            return '\n'.join(video_info)
        except Exception as e:
            return f"处理视频文件时出错: {str(e)}"
    
    def process_pdf_file(self, file_path: str) -> str:
        """处理PDF文件"""
        if not self.libraries['PyPDF2']:
            return "未安装PyPDF2库，无法处理PDF文件"
        
        try:
            with open(file_path, 'rb') as file:
                reader = self.libraries['PyPDF2'].PdfReader(file)
                full_text = []
                full_text.append(f"PDF文件：{os.path.basename(file_path)}")
                full_text.append(f"页数：{len(reader.pages)}")
                
                for page_num in range(len(reader.pages)):
                    page = reader.pages[page_num]
                    text = page.extract_text()
                    if text:
                        full_text.append(f"\n=== 第 {page_num + 1} 页 ===")
                        full_text.append(text)
                
                return '\n'.join(full_text)
        except Exception as e:
            return f"处理PDF文件时出错: {str(e)}"
    
    def process_wps_presentation(self, file_path: str) -> str:
        """处理WPS演示文稿文件（.dps）"""
        try:
            # WPS演示文稿文件本质上是压缩文件，尝试用zipfile打开
            from zipfile import ZipFile
            import xml.etree.ElementTree as ET
            
            full_text = []
            full_text.append(f"WPS演示文稿：{os.path.basename(file_path)}")
            
            with ZipFile(file_path, 'r') as zip_ref:
                # 检查是否包含PPTX格式的核心文件
                if 'ppt/slides/' in str(zip_ref.namelist()):
                    # 处理类似PPTX格式的WPS演示文稿
                    slide_files = [f for f in zip_ref.namelist() if f.startswith('ppt/slides/slide') and f.endswith('.xml')]
                    full_text.append(f"幻灯片数量：{len(slide_files)}")
                    
                    for slide_file in slide_files:
                        with zip_ref.open(slide_file) as xml_file:
                            tree = ET.parse(xml_file)
                            root = tree.getroot()
                            
                            # 定义命名空间
                            ns = {
                                'p': 'http://schemas.openxmlformats.org/presentationml/2006/main',
                                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
                            }
                            
                            # 提取幻灯片编号
                            slide_num = slide_file.split('slide')[1].split('.')[0]
                            full_text.append(f"\n=== 幻灯片 {slide_num} ===")
                            
                            # 查找所有文本元素
                            text_elements = root.findall('.//a:t', ns)
                            for text_elem in text_elements:
                                if text_elem.text:
                                    full_text.append(text_elem.text)
                else:
                    # 处理传统WPS演示文稿格式
                    return f"WPS演示文稿文件 {file_path} 使用了传统格式，暂时不支持直接解析。请将其另存为标准PPTX格式后重新上传。"
            
            if len(full_text) > 2:
                return '\n'.join(full_text)
            else:
                return f"成功解析WPS演示文稿文件，但未提取到文本内容。文件：{file_path}"
        except Exception as e:
            return f"处理WPS演示文稿文件时出错：{str(e)}。请将文件另存为标准PPTX格式后重新上传。"
    
    def extract_content(self, file_path: str) -> Dict[str, Any]:
        """提取文件内容的主方法"""
        file_name = os.path.basename(file_path)
        file_size = os.path.getsize(file_path)
        file_type = os.path.splitext(file_path)[1].lower()
        
        content = self.process_file(file_path)
        
        return {
            'file_name': file_name,
            'file_path': file_path,
            'file_size': file_size,
            'file_type': file_type,
            'content': content
        }